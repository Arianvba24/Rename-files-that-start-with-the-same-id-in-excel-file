import os
import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\Cash\Documents\pruebas_python\proyectos\os\main_file\main_file.xlsm")
ws = wb.active


remove_files= []
for i in range(2,1001):
    if ws[f"A{i}"].value is not None:
        remove_files.append(ws[f"A{i}"].value + ".xlsm")
        

last_files = {x for x in remove_files}

last_files = list(last_files)


values = list(filter(lambda x: x[-4:]=="xlsm",os.listdir(r"C:\Users\Cash\Documents\pruebas_python\proyectos\os")))

values

for value in values:
    if value in last_files:
        
        old_name = fr"C:\Users\Cash\Documents\pruebas_python\proyectos\os/{value}"
        new_name = fr"C:\Users\Cash\Documents\pruebas_python\proyectos\os/REMOVE_{value}"
        os.rename(old_name, new_name)
    else:
        pass
    
    

